import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill_azure = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")  # Azure blue
header_fill_aws = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")    # AWS orange
header_fill_cat = PatternFill(start_color="333333", end_color="333333", fill_type="solid")     # Dark gray
title_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
title_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
sub_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
light_blue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
light_orange = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

def style_row(ws, row, fills=None, font=None, alignment=None):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if font: cell.font = font
        if alignment: cell.alignment = alignment
        elif not alignment: cell.alignment = wrap_align
        if fills:
            if isinstance(fills, list):
                cell.fill = fills[col - 1] if col - 1 < len(fills) else fills[-1]
            else:
                cell.fill = fills

# ═══════════════════════════════════════════════════════════════════
# SHEET 1: COST COMPARISON
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Cost Comparison"
ws1.sheet_properties.tabColor = "0078D4"

# Column widths
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 25
ws1.column_dimensions["C"].width = 15
ws1.column_dimensions["D"].width = 25
ws1.column_dimensions["E"].width = 15
ws1.column_dimensions["F"].width = 20

# Title
ws1.merge_cells("A1:F1")
ws1["A1"] = "Solo E-Commerce — Azure vs AWS Cost Comparison (Monthly)"
ws1["A1"].font = title_font
ws1["A1"].fill = title_fill
ws1["A1"].alignment = center_align

# Headers
headers = ["Category", "Azure Service", "Azure Cost", "AWS Service", "AWS Cost", "Winner"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    fill = header_fill_azure if i in (2, 3) else header_fill_aws if i in (4, 5) else header_fill_cat
    cell.fill = fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Data rows
cost_data = [
    # Starter Tier
    ["── STARTER TIER (Low Traffic) ──", "", "", "", "", ""],
    ["Backend Hosting", "App Service B1 (1 vCPU, 1.75 GB)", "$13/mo", "Lightsail ($5) or EC2 t3.micro", "$5–8/mo", "AWS"],
    ["Database (PostgreSQL)", "Azure DB for PostgreSQL Flex B1ms", "$12/mo", "RDS db.t3.micro or Lightsail DB", "$10–15/mo", "Similar"],
    ["Frontend (Static)", "Azure Static Web Apps (Free)", "$0", "S3 + CloudFront", "$0.50–1/mo", "Azure"],
    ["File/Image Storage", "Azure Blob Storage (5 GB)", "$0.50/mo", "S3 (5 GB)", "$0.15/mo", "AWS"],
    ["SSL Certificate", "Included (App Service)", "$0", "ACM (Free with ALB/CF)", "$0", "Tie"],
    ["Domain / DNS", "Azure DNS", "$0.50/mo", "Route 53", "$0.50/mo", "Tie"],
    ["Monitoring", "Application Insights (Basic)", "$0", "CloudWatch (Basic)", "$0", "Tie"],
    ["TOTAL STARTER", "", "≈ $26–28/mo", "", "≈ $16–25/mo", "AWS ($8–10 cheaper)"],
    ["", "", "", "", "", ""],
    # Growth Tier
    ["── GROWTH TIER (Moderate Traffic) ──", "", "", "", "", ""],
    ["Backend Hosting", "App Service S1 (1 vCPU, 1.75 GB)", "$55/mo", "EC2 t3.small + ALB", "$25–35/mo", "AWS"],
    ["Database (PostgreSQL)", "Azure DB Flex B2s (2 vCPU, 4 GB)", "$50/mo", "RDS db.t3.small", "$35/mo", "AWS"],
    ["Frontend (Static)", "Azure Static Web Apps (Standard)", "$9/mo", "S3 + CloudFront", "$2–5/mo", "AWS"],
    ["File/Image Storage", "Azure Blob Storage (50 GB)", "$3/mo", "S3 (50 GB)", "$1.50/mo", "AWS"],
    ["CDN", "Azure CDN (50 GB transfer)", "$4/mo", "CloudFront (50 GB)", "$4/mo", "Tie"],
    ["Email (Transactional)", "SendGrid (via Azure Marketplace)", "$10–15/mo", "SES", "$1–3/mo", "AWS"],
    ["Monitoring", "Application Insights", "$5/mo", "CloudWatch", "$3/mo", "AWS"],
    ["TOTAL GROWTH", "", "≈ $136–141/mo", "", "≈ $72–86/mo", "AWS (≈$55 cheaper)"],
    ["", "", "", "", "", ""],
    # Production Tier
    ["── PRODUCTION TIER (High Traffic) ──", "", "", "", "", ""],
    ["Backend Hosting", "App Service P1v3 (2 vCPU, 8 GB)", "$110/mo", "EC2 t3.medium + ALB", "$55–65/mo", "AWS"],
    ["Database (PostgreSQL)", "Azure DB Flex D2s_v3 (2 vCPU, 8 GB)", "$100/mo", "RDS db.t3.medium (Multi-AZ)", "$80–100/mo", "AWS slightly"],
    ["Frontend (Static)", "Azure Static Web Apps (Standard)", "$9/mo", "S3 + CloudFront", "$5–10/mo", "Similar"],
    ["File/Image Storage", "Azure Blob Storage (200 GB)", "$10/mo", "S3 (200 GB)", "$5/mo", "AWS"],
    ["CDN", "Azure CDN (200 GB transfer)", "$14/mo", "CloudFront (200 GB)", "$15/mo", "Tie"],
    ["Redis Cache", "Azure Cache for Redis C0", "$16/mo", "ElastiCache t3.micro", "$13/mo", "AWS slightly"],
    ["Email", "SendGrid Standard", "$20/mo", "SES", "$5/mo", "AWS"],
    ["Backups", "Azure Backup", "$5/mo", "Automated RDS + S3", "$5/mo", "Tie"],
    ["Monitoring + Alerts", "Application Insights + Log Analytics", "$15/mo", "CloudWatch + X-Ray", "$10/mo", "AWS"],
    ["TOTAL PRODUCTION", "", "≈ $299–304/mo", "", "≈ $193–228/mo", "AWS (≈$80–100 cheaper)"],
]

row = 3
for data in cost_data:
    for col, val in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
    # Style section headers
    if data[0].startswith("──"):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).font = sub_font
            ws1.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    # Style totals
    if data[0].startswith("TOTAL"):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).font = Font(name="Calibri", bold=True, size=11)
            ws1.cell(row=row, column=col).fill = yellow_fill
    # Color Azure/AWS columns
    if not data[0].startswith("──") and not data[0].startswith("TOTAL") and data[0]:
        ws1.cell(row=row, column=2).fill = light_blue
        ws1.cell(row=row, column=3).fill = light_blue
        ws1.cell(row=row, column=4).fill = light_orange
        ws1.cell(row=row, column=5).fill = light_orange
        winner = data[5]
        if "AWS" in winner:
            ws1.cell(row=row, column=6).fill = light_orange
        elif "Azure" in winner:
            ws1.cell(row=row, column=6).fill = light_blue
        elif winner in ("Tie", "Similar"):
            ws1.cell(row=row, column=6).fill = green_fill
    row += 1

# ═══════════════════════════════════════════════════════════════════
# SHEET 2: FEATURE COMPARISON
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Feature Comparison")
ws2.sheet_properties.tabColor = "FF9900"

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 35
ws2.column_dimensions["D"].width = 20

ws2.merge_cells("A1:D1")
ws2["A1"] = "Solo E-Commerce — Azure vs AWS Feature Comparison"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2["A1"].alignment = center_align

feat_headers = ["Feature", "Azure", "AWS", "Winner"]
for i, h in enumerate(feat_headers, 1):
    cell = ws2.cell(row=2, column=i, value=h)
    fill = header_fill_azure if i == 2 else header_fill_aws if i == 3 else header_fill_cat
    cell.fill = fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

feat_data = [
    ["── EASE OF USE ──", "", "", ""],
    ["Initial Setup Complexity", "Very Easy — Portal or VS Code extension\nOne-click deploy from GitHub", "Moderate — More services to configure\nIAM roles, VPC, Security Groups", "Azure"],
    ["Learning Curve", "Gentle — Familiar UI, guided wizards\nGood for beginners", "Steeper — More concepts to learn\nBut better documentation", "Azure"],
    ["Deployment from GitHub", "GitHub Actions (built-in template)\nAzure DevOps integration", "GitHub Actions + CodeDeploy\nor Amplify for frontend", "Azure (simpler)"],
    ["CLI Tools", "Azure CLI (az) — clean & consistent", "AWS CLI (aws) — very powerful\nMore verbose", "Azure (simpler)"],
    ["VS Code Integration", "Excellent — First-party extensions\nDeploy right from editor", "Good — AWS Toolkit extension\nNot as integrated", "Azure"],
    ["", "", "", ""],
    ["── SCALABILITY ──", "", "", ""],
    ["Auto-Scaling", "App Service built-in autoscale\nEasy to configure rules", "EC2 Auto Scaling Groups\nMore flexible but more config", "AWS (more flexible)"],
    ["Global Reach", "60+ regions worldwide", "30+ regions worldwide\nMore edge locations for CDN", "AWS (more CDN edges)"],
    ["Serverless Option", "Azure Functions\n(can run NestJS via custom handler)", "AWS Lambda + API Gateway\n(mature, well-documented)", "AWS"],
    ["Container Option", "Azure Container Apps\n(simpler than AKS)", "ECS Fargate\n(no server management)", "Tie"],
    ["Database Scaling", "Flexible Server scales well\nRead replicas available", "RDS scales well\nAurora for extreme scale", "AWS (Aurora advantage)"],
    ["", "", "", ""],
    ["── RELIABILITY & SUPPORT ──", "", "", ""],
    ["SLA (Uptime)", "99.95% for App Service\n99.99% for zone-redundant", "99.99% for EC2\n99.95% for RDS Multi-AZ", "Tie"],
    ["Free Support", "Azure Community Support\nMSDN Forums, Stack Overflow", "AWS Community Support\nre:Post, Stack Overflow", "Tie"],
    ["Paid Support", "Starts at $29/mo (Developer)\n$100/mo (Standard)", "Starts at $29/mo (Developer)\n$100/mo (Business)", "Tie"],
    ["Incident Response", "15 min for Critical (Premier)", "15 min for Critical (Enterprise)", "Tie"],
    ["Status Dashboard", "Azure Status Page", "AWS Health Dashboard", "Tie"],
    ["", "", "", ""],
    ["── SECURITY ──", "", "", ""],
    ["SSL/TLS", "Free managed certificates\nAuto-renewal", "ACM free certificates\nAuto-renewal with ALB/CF", "Tie"],
    ["DDoS Protection", "Azure DDoS Protection Basic (free)\nStandard: $2,944/mo", "AWS Shield Standard (free)\nAdvanced: $3,000/mo", "Tie"],
    ["WAF (Web App Firewall)", "Azure WAF on App Gateway\n~$20-50/mo", "AWS WAF\n~$5-20/mo per ACL", "AWS (cheaper)"],
    ["Identity Management", "Azure AD / Entra ID\nExcellent enterprise SSO", "IAM — very granular\nCognito for user auth", "Azure (enterprise)\nAWS (granularity)"],
    ["Secrets Management", "Azure Key Vault\nFree for first 10K operations", "AWS Secrets Manager\n$0.40/secret/month", "Azure (cheaper)"],
    ["", "", "", ""],
    ["── SOLO-SPECIFIC FIT ──", "", "", ""],
    ["NestJS Backend", "App Service runs Node natively\nPM2 or built-in process mgmt", "EC2, Elastic Beanstalk,\nor Lightsail — all work well", "Azure (simpler setup)"],
    ["Flutter Web Frontend", "Static Web Apps — perfect fit\nFree tier generous", "S3 + CloudFront — battle-tested\nSlightly more setup", "Azure (easier)"],
    ["PostgreSQL Database", "Flexible Server — managed,\nauto-backups, point-in-time recovery", "RDS PostgreSQL — mature,\nmore instance size options", "AWS (more options)"],
    ["Image/File Upload", "Blob Storage — simple SDK\nCDN integration", "S3 — industry standard\nMore tooling available", "AWS"],
    ["Email (Order Confirm)", "SendGrid via Marketplace\nor Communication Services", "SES — very cheap\n$0.10 per 1000 emails", "AWS (much cheaper)"],
    ["Payment Integration", "No difference — Stripe/etc\nare cloud-agnostic", "No difference — Stripe/etc\nare cloud-agnostic", "Tie"],
]

row = 3
for data in feat_data:
    for col, val in enumerate(data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
    if data[0].startswith("──"):
        for col in range(1, 5):
            ws2.cell(row=row, column=col).font = sub_font
            ws2.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    elif data[0]:
        ws2.cell(row=row, column=2).fill = light_blue
        ws2.cell(row=row, column=3).fill = light_orange
        w = data[3]
        if "Azure" in w:
            ws2.cell(row=row, column=4).fill = light_blue
        elif "AWS" in w:
            ws2.cell(row=row, column=4).fill = light_orange
        elif w == "Tie":
            ws2.cell(row=row, column=4).fill = green_fill
    row += 1

# Adjust row heights for multi-line cells
for r in range(3, row):
    ws2.row_dimensions[r].height = 40

# ═══════════════════════════════════════════════════════════════════
# SHEET 3: FREE TIER COMPARISON
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Free Tier")
ws3.sheet_properties.tabColor = "00B050"

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 35
ws3.column_dimensions["D"].width = 20

ws3.merge_cells("A1:D1")
ws3["A1"] = "Solo E-Commerce — Free Tier Comparison"
ws3["A1"].font = title_font
ws3["A1"].fill = title_fill
ws3["A1"].alignment = center_align

free_headers = ["Service", "Azure Free Tier", "AWS Free Tier", "Notes"]
for i, h in enumerate(free_headers, 1):
    cell = ws3.cell(row=2, column=i, value=h)
    fill = header_fill_azure if i == 2 else header_fill_aws if i == 3 else header_fill_cat
    cell.fill = fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

free_data = [
    ["Compute", "App Service F1: 60 min/day CPU\n(not practical for production)\n12 months: B1 Linux free", "EC2 t2.micro: 750 hrs/mo\nfor 12 months\n(practical for light traffic)", "AWS better\nfor production use"],
    ["Database", "Azure DB Flex: 750 hrs Burstable B1ms\nfor 12 months\n32 GB storage", "RDS: 750 hrs db.t2.micro\nfor 12 months\n20 GB storage", "Azure gives\nmore storage"],
    ["Static Hosting", "Static Web Apps: 2 custom domains\n0.5 GB storage, 100 GB bandwidth\nALWAYS FREE", "S3: 5 GB Standard Storage\nCloudFront: 1 TB/mo transfer\nfor 12 months only", "Azure: always free\nAWS: 12 months"],
    ["Object Storage", "Blob Storage: 5 GB LRS\nfor 12 months", "S3: 5 GB Standard\nfor 12 months", "Same"],
    ["CDN", "Not included in free tier", "CloudFront: 1 TB/mo\nfor 12 months", "AWS better"],
    ["DNS", "Not included in free tier\n(~$0.50/mo)", "Route 53: NOT free\n(~$0.50/mo per zone)", "Both cost ~$0.50"],
    ["Monitoring", "Application Insights:\n1 GB/mo logs — ALWAYS FREE", "CloudWatch: 10 custom metrics,\n5 GB logs — ALWAYS FREE", "Both good"],
    ["Email", "Not included\n(use SendGrid free: 100/day)", "SES: 3,000 messages/mo\nfor 12 months", "AWS better"],
    ["SSL/TLS", "Free managed certificates\nALWAYS FREE", "ACM: Free certificates\nALWAYS FREE", "Same"],
    ["Key Vault / Secrets", "Key Vault: 10,000 operations\nALWAYS FREE", "Secrets Manager: NOT free\n($0.40/secret/mo)", "Azure better"],
    ["", "", "", ""],
    ["EFFECTIVE FREE\nPERIOD", "12 months for compute/DB\nSome services always free", "12 months for compute/DB\nSome services always free", "Similar"],
    ["AFTER FREE TIER\nEXPIRES", "Minimum ~$26–28/mo\nfor starter setup", "Minimum ~$16–25/mo\nfor starter setup", "AWS ~$8–10\ncheaper"],
]

row = 3
for data in free_data:
    for col, val in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
    if data[0]:
        ws3.cell(row=row, column=2).fill = light_blue
        ws3.cell(row=row, column=3).fill = light_orange
    if data[0].startswith("EFFECTIVE") or data[0].startswith("AFTER"):
        for col in range(1, 5):
            ws3.cell(row=row, column=col).font = Font(name="Calibri", bold=True, size=11)
            ws3.cell(row=row, column=col).fill = yellow_fill
    row += 1
    ws3.row_dimensions[row - 1].height = 55

# ═══════════════════════════════════════════════════════════════════
# SHEET 4: RECOMMENDATION
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Recommendation")
ws4.sheet_properties.tabColor = "B8860B"

ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 45
ws4.column_dimensions["C"].width = 45

ws4.merge_cells("A1:C1")
ws4["A1"] = "Solo E-Commerce — Final Recommendation"
ws4["A1"].font = title_font
ws4["A1"].fill = title_fill
ws4["A1"].alignment = center_align

# Summary scores
ws4.merge_cells("A3:C3")
ws4["A3"] = "SCORECARD SUMMARY"
ws4["A3"].font = Font(name="Calibri", bold=True, size=13)
ws4["A3"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ws4["A3"].alignment = center_align

scores = [
    ["Criteria", "Azure", "AWS"],
    ["Monthly Cost (Starter)", "6/10 — Higher baseline cost", "9/10 — Lowest cost option"],
    ["Monthly Cost (Growth)", "5/10 — Significantly more expensive", "8/10 — Much more affordable"],
    ["Monthly Cost (Production)", "5/10 — ~$300/mo", "8/10 — ~$200/mo"],
    ["Ease of Setup", "9/10 — Portal, VS Code, one-click", "6/10 — More manual configuration"],
    ["Learning Curve", "8/10 — Gentle for beginners", "5/10 — Steeper, many concepts"],
    ["VS Code Integration", "10/10 — First-party, seamless", "7/10 — Good toolkit available"],
    ["Scalability", "8/10 — Autoscale built-in", "9/10 — More flexible, more options"],
    ["Documentation", "7/10 — Good but sometimes scattered", "9/10 — Excellent, more examples"],
    ["Community & Ecosystem", "7/10 — Growing fast", "9/10 — Largest cloud ecosystem"],
    ["Enterprise Features", "9/10 — Entra ID, compliance", "8/10 — IAM, compliance"],
    ["NestJS + Flutter Fit", "9/10 — App Service + Static Web Apps", "7/10 — Works well, more assembly"],
    ["", "", ""],
    ["OVERALL SCORE", "7.5/10", "7.8/10"],
]

row = 4
for i, data in enumerate(scores):
    for col, val in enumerate(data, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border
        if i == 0:
            fill = header_fill_azure if col == 2 else header_fill_aws if col == 3 else header_fill_cat
            cell.fill = fill
            cell.font = header_font
            cell.alignment = center_align
        elif data[0] == "OVERALL SCORE":
            cell.font = Font(name="Calibri", bold=True, size=12)
            cell.fill = yellow_fill
        elif data[0]:
            cell.font = normal_font
            ws4.cell(row=row, column=2).fill = light_blue
            ws4.cell(row=row, column=3).fill = light_orange
    row += 1

row += 1
ws4.merge_cells(f"A{row}:C{row}")
ws4.cell(row=row, column=1, value="RECOMMENDATION").font = Font(name="Calibri", bold=True, size=13)
ws4.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ws4.cell(row=row, column=1).alignment = center_align
row += 1

recs = [
    ["Choose AZURE if:", "• You want the simplest setup experience\n• You prefer VS Code-first workflow\n• You already have Microsoft 365 / Entra ID\n• Enterprise compliance is a priority\n• You value convenience over cost savings\n• Budget: ~$28/mo starter, ~$140/mo growth", ""],
    ["Choose AWS if:", "• Cost is the primary concern\n• You want maximum flexibility\n• You plan to scale significantly\n• You want the cheapest email (SES)\n• You're comfortable with more configuration\n• Budget: ~$18/mo starter, ~$80/mo growth", ""],
    ["", "", ""],
    ["FOR SOLO E-COMMERCE\nSPECIFICALLY:", "AZURE is recommended for your case because:\n\n1. Simpler deployment — App Service + Static Web Apps is a natural fit for NestJS + Flutter Web\n\n2. VS Code integration — Deploy directly from your editor\n\n3. Free tier — 12 months free compute + always-free Static Web Apps\n\n4. Less DevOps work — You can focus on the product, not infrastructure\n\n5. The ~$8-10/mo premium over AWS is worth the time savings\n\nStart with free tier → move to B1 ($13/mo) + Flex B1ms ($12/mo) = $26/mo when ready", ""],
]

for data in recs:
    for col, val in enumerate(data, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
    if "AZURE" in data[0] or data[0].startswith("Choose AZURE"):
        ws4.cell(row=row, column=1).fill = light_blue
        ws4.cell(row=row, column=1).font = sub_font
        ws4.merge_cells(f"B{row}:C{row}")
    elif "AWS" in data[0] or data[0].startswith("Choose AWS"):
        ws4.cell(row=row, column=1).fill = light_orange
        ws4.cell(row=row, column=1).font = sub_font
        ws4.merge_cells(f"B{row}:C{row}")
    elif data[0].startswith("FOR SOLO"):
        ws4.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
        ws4.cell(row=row, column=1).fill = green_fill
        ws4.merge_cells(f"B{row}:C{row}")
        ws4.cell(row=row, column=2).fill = green_fill
        ws4.row_dimensions[row].height = 140
    row += 1

# Auto-set some row heights
for r in range(4, row):
    if ws4.row_dimensions[r].height is None or ws4.row_dimensions[r].height < 35:
        ws4.row_dimensions[r].height = 35

# ═══════════════════════════════════════════════════════════════════
# SHEET 5: DEPLOYMENT ARCHITECTURE
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Deployment Architecture")
ws5.sheet_properties.tabColor = "7030A0"

ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 30
ws5.column_dimensions["C"].width = 35
ws5.column_dimensions["D"].width = 20

ws5.merge_cells("A1:D1")
ws5["A1"] = "Solo E-Commerce — Deployment Architecture"
ws5["A1"].font = title_font
ws5["A1"].fill = title_fill
ws5["A1"].alignment = center_align

arch_headers = ["Component", "Azure Architecture", "AWS Architecture", "Purpose"]
for i, h in enumerate(arch_headers, 1):
    cell = ws5.cell(row=2, column=i, value=h)
    fill = header_fill_azure if i == 2 else header_fill_aws if i == 3 else header_fill_cat
    cell.fill = fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

arch_data = [
    ["Flutter Web (Frontend)", "Azure Static Web Apps\n→ Auto-deploy from GitHub\n→ Global CDN built-in\n→ Custom domain + SSL", "S3 Static Website\n→ CloudFront CDN in front\n→ Route 53 for DNS\n→ ACM for SSL", "Serves Flutter web\nbuild output (HTML/JS/CSS)"],
    ["NestJS API (Backend)", "Azure App Service (Linux)\n→ Node.js 18+ runtime\n→ GitHub Actions CI/CD\n→ App Settings for env vars", "EC2 instance or\nElastic Beanstalk\n→ PM2 process manager\n→ CodeDeploy CI/CD\n→ Environment properties", "Runs NestJS API server\non port 3000"],
    ["PostgreSQL Database", "Azure Database for PostgreSQL\nFlexible Server\n→ Automated backups\n→ Point-in-time restore\n→ VNet integration", "Amazon RDS for PostgreSQL\n→ Automated backups\n→ Point-in-time restore\n→ VPC security groups\n→ Multi-AZ option", "Stores all application\ndata (products, users,\norders, etc.)"],
    ["File/Image Storage", "Azure Blob Storage\n→ Container for uploads\n→ CDN endpoint\n→ SAS tokens for access", "Amazon S3\n→ Bucket for uploads\n→ CloudFront distribution\n→ Presigned URLs", "Stores product images,\nbanner images,\nuser uploads"],
    ["Email Service", "SendGrid (Azure Marketplace)\nor Azure Communication Services\n→ Order confirmations\n→ Password reset emails", "Amazon SES\n→ Order confirmations\n→ Password reset emails\n→ Very cost-effective\n($0.10/1000 emails)", "Transactional emails\nfor order flow"],
    ["Caching (Optional)", "Azure Cache for Redis\n→ Session storage\n→ API response cache", "Amazon ElastiCache (Redis)\n→ Session storage\n→ API response cache", "Performance boost\nfor high traffic"],
    ["Monitoring", "Application Insights\n→ Request tracing\n→ Error tracking\n→ Performance metrics", "CloudWatch + X-Ray\n→ Metrics & alarms\n→ Request tracing\n→ Log aggregation", "Monitor app health,\nperformance, errors"],
    ["CI/CD Pipeline", "GitHub Actions\n→ Build → Test → Deploy\n→ Azure CLI in workflow\n→ Slot deployments", "GitHub Actions\n→ Build → Test → Deploy\n→ AWS CLI in workflow\n→ Blue/green with CodeDeploy", "Automated build and\ndeployment pipeline"],
]

row = 3
for data in arch_data:
    for col, val in enumerate(data, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border
    ws5.cell(row=row, column=1).font = sub_font
    ws5.cell(row=row, column=2).fill = light_blue
    ws5.cell(row=row, column=3).fill = light_orange
    ws5.row_dimensions[row].height = 80
    row += 1

# ── Save ──
output_path = r"D:\Solo Website\docs\Azure_vs_AWS_Comparison.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
